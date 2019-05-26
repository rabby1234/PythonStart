import csv

import openpyxl
import glob
import shutil
import os
from datetime import datetime, timedelta

intial_path = "/home/rabby/practise/files"
sub_folders = ["test1", "test2", "test3"]
destination_path = "/home/rabby/practise/destination"
file_path = "/home/rabby/practise/folders.csv"
extension ="xlsx"
duration = 10


def scan_folder(path, pattern="", extension=""):
    print("Started to scan path [ {} ] ".format(path))
    file_list = glob.iglob(path+"/*{}*{}".format(pattern, extension))
    for file in sorted(file_list, key=os.path.getctime, reverse=True):
        if created_in_duration(file):
            shutil.copy2(file, destination_path)
            print("{} - has been copied to {}".format(file, destination_path))
        else:
            print("Done scanning all files in folder {} ".format(path))
            break

def created_in_duration(file_name):
    file_mod_time = datetime.fromtimestamp(os.stat(file_name).st_ctime)  # This is a datetime.datetime object!
    now = datetime.today()
    interval = timedelta(minutes=duration)
    if now - file_mod_time > interval:
        print("its not created in last {} minutes ".format(interval.seconds/60))
        return False
    else:
        print("its okay to move")
        return True

def read_excel(path,foldername_column_index,pattern_column_index):
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    # Will print a particular row value
    output = list()
    for i in range(1, max_row + 1):
        folder_name_obj = sheet_obj.cell(row=i, column=foldername_column_index)
        pattern_obj = sheet_obj.cell(row=i, column=pattern_column_index)
        output.append((folder_name_obj.value, pattern_obj.value))
    return output

def read_file(filename):
    output = list()
    with open(filename, 'r') as csvfile:
        # creating a csv reader object
        csvreader = csv.reader(csvfile)
        # extracting each data row one by one
        for row in csvreader:
            output.append(tuple([col for col in row]))
    return output

if __name__ == "__main__":
    if os.path.exists(intial_path):
        sub_folders = read_file(file_path) #read_excel(file_path,1,2)
        for folder, pattern in sub_folders:
            folder_path = os.path.join(intial_path, folder)
            if os.path.exists(folder_path):
                scan_folder(folder_path, pattern=pattern, extension=extension)
            else:
                print("[{}] does not exist".format(folder_path))
    else:
        print("[ {} ] Path does not exist".format(intial_path))
